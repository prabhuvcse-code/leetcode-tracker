"""
LeetCode Daily Tracker v4
- Reads Name + LeetCode ID from Sheet1 of the Excel file
- Updates Daily Tracking, Overall Stats, and Contest sheets
- Sends daily email report at 9:30 PM with Excel attached
- Run via RUN_TRACKER.bat
"""

import os
import sys
import time
import json
import smtplib
import subprocess
import requests
import openpyxl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
from pathlib import Path

# ─────────────────────────── CONFIGURATION ───────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent.resolve()
EXCEL_FILE   = SCRIPT_DIR / "leetcode_OVERALL_DAILY_REPORT.xlsx"
LOG_FILE     = SCRIPT_DIR / "tracker_log.txt"
TASK_NAME_FETCH = "LeetCodeDailyTracker_Fetch"
TASK_NAME_EMAIL = "LeetCodeDailyTracker_Email"

# ──────────────── EMAIL CONFIGURATION ────────────────────────────────────────
import os
SENDER_EMAIL    = os.environ.get("SENDER_EMAIL", "")
SENDER_PASSWORD = os.environ.get("SENDER_PASSWORD", "")
```

Save the file and re-upload it to GitHub.

---

## ▶️ Step 6 — Test It Right Now (Don't Wait Till Night)

1. In your repo → click **"Actions"** tab
2. Click **"LeetCode Daily Tracker"** on the left
3. Click **"Run workflow"** → **"Run workflow"** (green button)
4. Watch it run live — takes 5-10 minutes depending on student count
5. Check `prabhuv.cse@citchennai.net` for the email! ✅

---

## ✅ What Happens Every Night
```
9:15 PM IST  →  GitHub servers wake up automatically
               →  Fetches all student LeetCode data
               →  Updates Excel file
               →  Sends email to prabhuv.cse@citchennai.net
               →  Saves updated Excel back to repo
               →  Done! Your PC can be completely OFF 🎉
# ─────────────────────────── LOGGING ─────────────────────────────────────────
def log(msg=""):
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"{ts}  {msg}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")

# ─────────────────────────── LEETCODE API ────────────────────────────────────
HEADERS = {
    "Content-Type": "application/json",
    "Referer":      "https://leetcode.com",
    "User-Agent":   "Mozilla/5.0",
}
API = "https://leetcode.com/graphql"

def gql(query: str, variables: dict, timeout=15):
    try:
        r = requests.post(API, json={"query": query, "variables": variables},
                          headers=HEADERS, timeout=timeout)
        r.raise_for_status()
        return r.json().get("data", {})
    except Exception as e:
        return None

def fetch_overall_stats(username: str):
    q = """
    query userProblemsSolved($username: String!) {
      matchedUser(username: $username) {
        submitStatsGlobal {
          acSubmissionNum {
            difficulty count
          }
        }
      }
    }"""
    data = gql(q, {"username": username.strip()})
    if not data:
        return None
    try:
        nums = data["matchedUser"]["submitStatsGlobal"]["acSubmissionNum"]
        d = {x["difficulty"]: x["count"] for x in nums}
        return {
            "easy":   d.get("Easy",  0),
            "medium": d.get("Medium",0),
            "hard":   d.get("Hard",  0),
            "total":  d.get("All",   0),
        }
    except Exception:
        return None

def fetch_daily_stats(username: str, target_date: date):
    q = """
    query recentAcSubmissions($username: String!, $limit: Int!) {
      recentAcSubmissionList(username: $username, limit: $limit) {
        id
        timestamp
        title
        titleSlug
      }
    }"""
    data = gql(q, {"username": username.strip(), "limit": 50})
    if not data:
        return None
    try:
        subs = data.get("recentAcSubmissionList") or []
    except Exception:
        return None

    seen = {}
    for s in subs:
        ts    = int(s.get("timestamp", 0))
        sdate = date.fromtimestamp(ts)
        if sdate == target_date:
            slug = s["titleSlug"]
            if slug not in seen:
                seen[slug] = True

    if not seen:
        return {"easy": 0, "medium": 0, "hard": 0, "total": 0}

    diff_q = """
    query questionDifficulty($titleSlug: String!) {
      question(titleSlug: $titleSlug) {
        difficulty
      }
    }"""
    counts = {"Easy": 0, "Medium": 0, "Hard": 0}
    for slug in seen:
        d2 = gql(diff_q, {"titleSlug": slug}, timeout=10)
        try:
            diff = d2["question"]["difficulty"]
            counts[diff] = counts.get(diff, 0) + 1
        except Exception:
            counts["Easy"] += 1
        time.sleep(0.2)

    return {
        "easy":   counts["Easy"],
        "medium": counts["Medium"],
        "hard":   counts["Hard"],
        "total":  counts["Easy"] + counts["Medium"] + counts["Hard"],
    }

def fetch_contest_stats(username: str):
    q = """
    query userContestRanking($username: String!) {
      userContestRanking(username: $username) {
        rating
        globalRanking
        attendedContestsCount
      }
      userContestRankingHistory(username: $username) {
        attended
        rating
        ranking
        problemsSolved
        totalProblems
        contest { title startTime }
      }
    }"""
    data = gql(q, {"username": username.strip()})
    if not data:
        return None
    try:
        cr   = data.get("userContestRanking") or {}
        hist = data.get("userContestRankingHistory") or []
        attended = [h for h in hist if h.get("attended")]
        last = attended[-1] if attended else None
        return {
            "rating":          round(cr.get("rating", 0), 1) if cr.get("rating") else "N/A",
            "global_ranking":  cr.get("globalRanking", "N/A"),
            "contests_count":  cr.get("attendedContestsCount", 0),
            "last_contest":    last["contest"]["title"] if last else "N/A",
            "last_attended":   "Yes" if last else "No",
            "last_ranking":    last.get("ranking", "N/A") if last else "N/A",
            "last_solved":     f"{last.get('problemsSolved','?')}/{last.get('totalProblems','?')}" if last else "N/A",
        }
    except Exception:
        return None

# ─────────────────────────── EXCEL HELPERS ───────────────────────────────────
HDR_FILL   = PatternFill("solid", start_color="1F3864", end_color="1F3864")
DAY_FILL   = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
ALT_FILL   = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
WHT_FILL   = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DAY_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
NORM_FONT  = Font(name="Arial", size=10)
CTR        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT       = Alignment(horizontal="left",   vertical="center")

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def style(cell, font=None, fill=None, align=None, num_fmt=None):
    if font:    cell.font      = font
    if fill:    cell.fill      = fill
    if align:   cell.alignment = align
    if num_fmt: cell.number_format = num_fmt
    cell.border = thin_border()

# ─────────────────────────── READ SHEET1 ─────────────────────────────────────
def read_students(wb) -> list[dict]:
    ws   = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    header_row = 0
    for i, row in enumerate(rows):
        cells = [str(c).strip().lower() if c else "" for c in row]
        if "name" in cells and any("leetcode" in c for c in cells):
            header_row = i
            break
    students = []
    for row in rows[header_row + 1:]:
        if not any(row):
            continue
        name  = str(row[0]).strip() if row[0] else ""
        reg   = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        lc_id = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        if name and lc_id and lc_id.lower() != "nan":
            students.append({"name": name, "reg": reg, "lc_id": lc_id})
    return students

# ─────────────────────────── DAILY TRACKING SHEET ────────────────────────────
def update_daily_sheet(wb, students: list[dict], today: date, daily_data: dict):
    ws = wb["Daily Tracking"]
    all_vals = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    row0 = list(all_vals[0]) if all_vals else []
    row1 = list(all_vals[1]) if len(all_vals) > 1 else []
    today_str = str(today)

    date_col_map = {}
    col = 3
    for c in range(2, len(row0)):
        v = row0[c]
        if v and str(v).strip():
            date_col_map[str(v).strip()] = c + 1
            col = c + 1

    if today_str not in date_col_map:
        next_col = (max(date_col_map.values()) + 4) if date_col_map else 3
        date_col_map[today_str] = next_col

        ws.merge_cells(start_row=1, start_column=next_col,
                       end_row=1,   end_column=next_col + 3)
        dc = ws.cell(row=1, column=next_col, value=today_str)
        style(dc, font=DAY_FONT, fill=DAY_FILL, align=CTR)

        for i, sub in enumerate(["Easy", "Medium", "Hard", "Total"]):
            sc = ws.cell(row=2, column=next_col + i, value=sub)
            style(sc, font=HDR_FONT, fill=HDR_FILL, align=CTR)
            ws.column_dimensions[get_column_letter(next_col + i)].width = 8

    start_col = date_col_map[today_str]

    name_row_map = {}
    for r in ws.iter_rows(min_row=3):
        v = r[0].value
        if v:
            name_row_map[str(v).strip()] = r[0].row

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 16

    if not ws.cell(1, 1).value:
        h1 = ws.cell(1, 1, "Name");        style(h1, font=HDR_FONT, fill=HDR_FILL, align=LEFT)
        h2 = ws.cell(2, 1, "Name");        style(h2, font=HDR_FONT, fill=HDR_FILL, align=LEFT)
        h3 = ws.cell(1, 2, "LeetCode ID"); style(h3, font=HDR_FONT, fill=HDR_FILL, align=CTR)
        h4 = ws.cell(2, 2, "LeetCode ID"); style(h4, font=HDR_FONT, fill=HDR_FILL, align=CTR)
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 22

    for idx, s in enumerate(students):
        name  = s["name"]
        lc_id = s["lc_id"]
        row   = name_row_map.get(name)
        if not row:
            row = ws.max_row + 1
            name_row_map[name] = row

        fill = ALT_FILL if idx % 2 == 0 else WHT_FILL
        nc = ws.cell(row=row, column=1, value=name);  style(nc, font=NORM_FONT, fill=fill, align=LEFT)
        ic = ws.cell(row=row, column=2, value=lc_id); style(ic, font=NORM_FONT, fill=fill, align=CTR)

        dd = daily_data.get(lc_id, {})
        for i, key in enumerate(["easy", "medium", "hard", "total"]):
            val = dd.get(key, 0) if dd else 0
            fc  = ws.cell(row=row, column=start_col + i, value=val)
            style(fc, font=NORM_FONT, fill=fill, align=CTR)

# ─────────────────────────── OVERALL STATS SHEET ─────────────────────────────
def update_overall_sheet(wb, students: list[dict], today: date, overall_data: dict):
    ws = wb["Overall Stats"]
    ws.delete_rows(1, ws.max_row)
    today_str = str(today)

    ws.merge_cells("C1:F1")
    ws.cell(1, 1, "Name");          style(ws.cell(1,1), HDR_FONT, HDR_FILL, LEFT)
    ws.cell(2, 1, "Name");          style(ws.cell(2,1), HDR_FONT, HDR_FILL, LEFT)
    ws.cell(1, 2, "LeetCode ID");   style(ws.cell(1,2), HDR_FONT, HDR_FILL, CTR)
    ws.cell(2, 2, "LeetCode ID");   style(ws.cell(2,2), HDR_FONT, HDR_FILL, CTR)
    h = ws.cell(1, 3, f"Overall Stats (as of {today_str})")
    style(h, DAY_FONT, DAY_FILL, CTR)
    for i, sub in enumerate(["Easy", "Medium", "Hard", "Total"]):
        sc = ws.cell(2, 3 + i, sub); style(sc, HDR_FONT, HDR_FILL, CTR)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    for col in ["C","D","E","F"]: ws.column_dimensions[col].width = 9

    for idx, s in enumerate(students):
        row   = idx + 3
        fill  = ALT_FILL if idx % 2 == 0 else WHT_FILL
        lc_id = s["lc_id"]
        od    = overall_data.get(lc_id, {})

        ws.cell(row, 1, s["name"]);  style(ws.cell(row,1), NORM_FONT, fill, LEFT)
        ws.cell(row, 2, lc_id);      style(ws.cell(row,2), NORM_FONT, fill, CTR)
        for i, key in enumerate(["easy","medium","hard","total"]):
            val = od.get(key, "?") if od else "?"
            c = ws.cell(row, 3+i, val); style(c, NORM_FONT, fill, CTR)

# ─────────────────────────── CONTEST SHEET ───────────────────────────────────
def update_contest_sheet(wb, students: list[dict], contest_data: dict):
    ws = wb["Contest"]
    ws.delete_rows(1, ws.max_row)

    cols = ["Name","LeetCode ID","Contest Rating","Global Ranking",
            "Contests Attended","Last Contest","Attended?","Last Ranking","Problems Solved"]
    widths = [22, 22, 14, 14, 17, 22, 10, 14, 15]

    for i, (col, w) in enumerate(zip(cols, widths)):
        c = ws.cell(1, i+1, col); style(c, HDR_FONT, HDR_FILL, CTR)
        ws.column_dimensions[get_column_letter(i+1)].width = w

    for idx, s in enumerate(students):
        row   = idx + 2
        fill  = ALT_FILL if idx % 2 == 0 else WHT_FILL
        lc_id = s["lc_id"]
        cd    = contest_data.get(lc_id, {})

        vals = [
            s["name"], lc_id,
            cd.get("rating",         "N/A"),
            cd.get("global_ranking", "N/A"),
            cd.get("contests_count", "N/A"),
            cd.get("last_contest",   "N/A"),
            cd.get("last_attended",  "N/A"),
            cd.get("last_ranking",   "N/A"),
            cd.get("last_solved",    "N/A"),
        ]
        for i, val in enumerate(vals):
            c = ws.cell(row, i+1, val)
            style(c, NORM_FONT, fill, CTR if i > 0 else LEFT)

# ─────────────────────────── EMAIL FUNCTION ──────────────────────────────────
def build_email_html(students, daily_data, today):
    """Build an HTML summary table split by class."""
    today_str = today.strftime("%d %B %Y")

    # Group students by class (based on reg number prefix or name — adapt as needed)
    # If Sheet1 has a 'class' column (index 3), use that; otherwise group all together
    # For now, we group by the 4th character of reg number or just show all in one table
    rows_html = ""
    total_solved_today = 0
    for idx, s in enumerate(students):
        lc_id = s["lc_id"]
        dd    = daily_data.get(lc_id, {})
        e, m, h, t = dd.get("easy",0), dd.get("medium",0), dd.get("hard",0), dd.get("total",0)
        total_solved_today += t
        bg = "#EAF0FB" if idx % 2 == 0 else "#FFFFFF"
        highlight = ' style="font-weight:bold; color:#1F6E43;"' if t > 0 else ""
        rows_html += f"""
        <tr style="background:{bg};">
          <td style="padding:6px 10px;">{idx+1}</td>
          <td style="padding:6px 10px;">{s['name']}</td>
          <td style="padding:6px 10px;">{s.get('reg','')}</td>
          <td style="padding:6px 10px;">{lc_id}</td>
          <td style="padding:6px 10px; text-align:center;">{e}</td>
          <td style="padding:6px 10px; text-align:center;">{m}</td>
          <td style="padding:6px 10px; text-align:center;">{h}</td>
          <td style="padding:6px 10px; text-align:center;"{highlight}>{t}</td>
        </tr>"""

    active_count = sum(1 for s in students if daily_data.get(s["lc_id"], {}).get("total", 0) > 0)

    html = f"""
    <html><body style="font-family:Arial,sans-serif; color:#222;">
    <h2 style="color:#1F3864;">📊 LeetCode Daily Report — {today_str}</h2>
    <p>
      <strong>Total Students:</strong> {len(students)} &nbsp;|&nbsp;
      <strong>Active Today:</strong> {active_count} &nbsp;|&nbsp;
      <strong>Problems Solved Today:</strong> {total_solved_today}
    </p>
    <table border="0" cellspacing="0" cellpadding="0"
           style="border-collapse:collapse; width:100%; font-size:13px;">
      <thead>
        <tr style="background:#1F3864; color:#fff;">
          <th style="padding:8px 10px;">#</th>
          <th style="padding:8px 10px; text-align:left;">Name</th>
          <th style="padding:8px 10px; text-align:left;">Reg No</th>
          <th style="padding:8px 10px; text-align:left;">LeetCode ID</th>
          <th style="padding:8px 10px;">Easy</th>
          <th style="padding:8px 10px;">Medium</th>
          <th style="padding:8px 10px;">Hard</th>
          <th style="padding:8px 10px;">Total</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
    <p style="margin-top:20px; color:#555; font-size:12px;">
      Auto-generated by LeetCode Daily Tracker • Full report attached as Excel file.
    </p>
    </body></html>"""
    return html

def send_email(today, students, daily_data):
    """Send daily report email with Excel attached."""
    log("  Preparing email...")

    if SENDER_EMAIL == "YOUR_GMAIL@gmail.com":
        log("  ⚠️  Email skipped: Please update SENDER_EMAIL and SENDER_PASSWORD in the script!")
        return

    today_str = today.strftime("%d %B %Y")
    subject   = f"LeetCode Daily Report — {today_str}"
    html_body = build_email_html(students, daily_data, today)

    msg = MIMEMultipart("mixed")
    msg["From"]    = SENDER_EMAIL
    msg["To"]      = RECIPIENT_EMAIL
    msg["Subject"] = subject

    msg.attach(MIMEText(html_body, "html"))

    # Attach Excel file
    if EXCEL_FILE.exists():
        with open(EXCEL_FILE, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition",
                        f'attachment; filename="{EXCEL_FILE.name}"')
        msg.attach(part)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            server.sendmail(SENDER_EMAIL, RECIPIENT_EMAIL, msg.as_string())
        log(f"  ✅ Email sent to {RECIPIENT_EMAIL}")
    except Exception as e:
        log(f"  ❌ Email failed: {e}")
        log(f"     Check SENDER_EMAIL, SENDER_PASSWORD, and Gmail App Password settings.")

# ─────────────────────────── TASK SCHEDULER ──────────────────────────────────
def register_tasks():
    py = sys.executable

    # Task 1: Fetch at 9:15 PM — runs the main tracker with --fetch flag
    cmd1 = (
        f'schtasks /Create /TN "{TASK_NAME_FETCH}" '
        f'/TR "\\"{py}\\" \\"{__file__}\\" --fetch" '
        f'/SC DAILY /ST 21:15 /F /RL HIGHEST'
    )
    r1 = subprocess.run(cmd1, shell=True, capture_output=True, text=True)
    if r1.returncode == 0:
        log("  ✅ Task 1: Fetch data at 9:15 PM registered")
    else:
        log("  ⚠️  Task 1 registration failed — run as Administrator once")

    # Task 2: Send email at 9:30 PM — runs with --email flag
    cmd2 = (
        f'schtasks /Create /TN "{TASK_NAME_EMAIL}" '
        f'/TR "\\"{py}\\" \\"{__file__}\\" --email" '
        f'/SC DAILY /ST 21:30 /F /RL HIGHEST'
    )
    r2 = subprocess.run(cmd2, shell=True, capture_output=True, text=True)
    if r2.returncode == 0:
        log("  ✅ Task 2: Send email at 9:30 PM registered")
    else:
        log("  ⚠️  Task 2 registration failed — run as Administrator once")

# ─────────────────────────── FETCH MODE ──────────────────────────────────────
def run_fetch():
    """Fetch LeetCode data and save to Excel. Called at 9:15 PM."""
    today = date.today()

    log()
    log("╔══════════════════════════════════════════════╗")
    log("║    LeetCode Daily Tracker v4 — FETCH         ║")
    log("╚══════════════════════════════════════════════╝")
    log(f"  Date   : {today}")
    log(f"  Excel  : {EXCEL_FILE}")

    if not EXCEL_FILE.exists():
        log(f"  ERROR: Excel file not found: {EXCEL_FILE}")
        return None, None

    wb       = openpyxl.load_workbook(EXCEL_FILE)
    students = read_students(wb)
    log(f"  Found {len(students)} students.\n")

    for sheet in ["Daily Tracking", "Overall Stats", "Contest"]:
        if sheet not in wb.sheetnames:
            wb.create_sheet(sheet)

    daily_data   = {}
    overall_data = {}
    contest_data = {}

    for s in students:
        name  = s["name"]
        lc_id = s["lc_id"].strip()
        log(f"  Fetching: {name}  ({lc_id})")

        try:
            dd = fetch_daily_stats(lc_id, today)
            daily_data[lc_id] = dd or {"easy":0,"medium":0,"hard":0,"total":0}
            d = daily_data[lc_id]
            log(f"    Daily  : E={d['easy']} M={d['medium']} H={d['hard']} T={d['total']}")
        except Exception as e:
            log(f"    [WARN] fetch_daily_stats: {e}")
            daily_data[lc_id] = {"easy":0,"medium":0,"hard":0,"total":0}

        try:
            od = fetch_overall_stats(lc_id)
            overall_data[lc_id] = od or {}
        except Exception as e:
            log(f"    [WARN] fetch_overall_stats: {e}")
            overall_data[lc_id] = {}

        try:
            cd = fetch_contest_stats(lc_id)
            contest_data[lc_id] = cd or {}
        except Exception as e:
            log(f"    [WARN] fetch_contest_stats: {e}")
            contest_data[lc_id] = {}

        time.sleep(1.0)

    log()
    log("  Updating Excel sheets...")
    update_daily_sheet(wb, students, today, daily_data)
    update_overall_sheet(wb, students, today, overall_data)
    update_contest_sheet(wb, students, contest_data)
    wb.save(EXCEL_FILE)
    log(f"  💾 Saved: {EXCEL_FILE}")

    # Save daily_data to a temp JSON for the email task to read
    cache_file = SCRIPT_DIR / "daily_cache.json"
    with open(cache_file, "w", encoding="utf-8") as f:
        json.dump({
            "date": str(today),
            "daily_data": daily_data,
            "students": students
        }, f, ensure_ascii=False, indent=2)
    log(f"  💾 Data cached for email: {cache_file}")
    log()

# ─────────────────────────── EMAIL MODE ──────────────────────────────────────
def run_email():
    """Read cached data and send email. Called at 9:30 PM."""
    today = date.today()

    log()
    log("╔══════════════════════════════════════════════╗")
    log("║    LeetCode Daily Tracker v4 — EMAIL         ║")
    log("╚══════════════════════════════════════════════╝")

    cache_file = SCRIPT_DIR / "daily_cache.json"
    if not cache_file.exists():
        log("  ⚠️  No cached data found. Running fetch first...")
        run_fetch()

    try:
        with open(cache_file, "r", encoding="utf-8") as f:
            cache = json.load(f)
        students   = cache["students"]
        daily_data = cache["daily_data"]
        cached_date = cache.get("date", "")
        if cached_date != str(today):
            log(f"  ⚠️  Cache is from {cached_date}, today is {today}. Re-fetching...")
            run_fetch()
            with open(cache_file, "r", encoding="utf-8") as f:
                cache = json.load(f)
            students   = cache["students"]
            daily_data = cache["daily_data"]
    except Exception as e:
        log(f"  ERROR reading cache: {e}")
        return

    send_email(today, students, daily_data)
    log()

# ─────────────────────────── FULL RUN (fetch + email together) ───────────────
def run_all():
    """Fetch ALL student data first, save Excel, then send email. One task does everything."""
    today = date.today()

    log()
    log("╔══════════════════════════════════════════════╗")
    log("║   LeetCode Daily Tracker v4 — FULL RUN       ║")
    log("╚══════════════════════════════════════════════╝")
    log(f"  Date   : {today}")
    log(f"  Excel  : {EXCEL_FILE}")
    log()

    if not EXCEL_FILE.exists():
        log(f"  ERROR: Excel file not found: {EXCEL_FILE}")
        return

    wb       = openpyxl.load_workbook(EXCEL_FILE)
    students = read_students(wb)
    log(f"  Found {len(students)} students.")
    log(f"  Step 1 of 2 — Fetching all LeetCode data...")
    log("=" * 55)

    for sheet in ["Daily Tracking", "Overall Stats", "Contest"]:
        if sheet not in wb.sheetnames:
            wb.create_sheet(sheet)

    daily_data   = {}
    overall_data = {}
    contest_data = {}

    for idx, s in enumerate(students):
        name  = s["name"]
        lc_id = s["lc_id"].strip()
        log(f"  [{idx+1}/{len(students)}] {name}  ({lc_id})")

        try:
            dd = fetch_daily_stats(lc_id, today)
            daily_data[lc_id] = dd or {"easy":0,"medium":0,"hard":0,"total":0}
            d = daily_data[lc_id]
            log(f"    Daily  : E={d['easy']} M={d['medium']} H={d['hard']} T={d['total']}")
        except Exception as e:
            log(f"    [WARN] fetch_daily_stats: {e}")
            daily_data[lc_id] = {"easy":0,"medium":0,"hard":0,"total":0}

        try:
            od = fetch_overall_stats(lc_id)
            overall_data[lc_id] = od or {}
        except Exception as e:
            log(f"    [WARN] fetch_overall_stats: {e}")
            overall_data[lc_id] = {}

        try:
            cd = fetch_contest_stats(lc_id)
            contest_data[lc_id] = cd or {}
        except Exception as e:
            log(f"    [WARN] fetch_contest_stats: {e}")
            contest_data[lc_id] = {}

        time.sleep(1.0)

    log()
    log("  ✅ All students fetched!")
    log("  Updating Excel sheets...")
    update_daily_sheet(wb, students, today, daily_data)
    update_overall_sheet(wb, students, today, overall_data)
    update_contest_sheet(wb, students, contest_data)
    wb.save(EXCEL_FILE)
    log(f"  💾 Excel saved: {EXCEL_FILE}")
    log()
    log("=" * 55)
    log("  Step 2 of 2 — Sending email now...")
    log("=" * 55)
    send_email(today, students, daily_data)
    log()
    log("  🎉 All done! Data fetched + Email sent.")
    log()

# ─────────────────────────── MAIN ────────────────────────────────────────────
def main():
    args = sys.argv[1:]

    if "--fetch" in args:
        run_fetch()
        return

    if "--email" in args:
        run_email()
        return

    # --run or default: fetch ALL data first, then send email
    run_all()

if __name__ == "__main__":
    main()
